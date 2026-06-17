# -*- coding: utf-8 -*-
"""Extract capability_map.xlsx into site/data.js (self-contained, no fetch needed)."""
import json, os, re
from collections import Counter, defaultdict
import openpyxl

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
XLSX = os.path.join(ROOT, "capability_map.xlsx")
OUT  = os.path.join(ROOT, "site", "data.js")

wb = openpyxl.load_workbook(XLSX)
cap_ws = wb["Capabilities"]
pr_ws  = wb["Prompts"]

cap_headers = [c.value for c in cap_ws[1]]
caps = []
for row in cap_ws.iter_rows(min_row=2, values_only=True):
    d = dict(zip(cap_headers, row))
    caps.append({
        "id": d["ID"],
        "area": d["Capability area"],
        "desc": d["Description"],
        "layer": (d["Layer"] or "").strip(),
        "quiverData": d["Quiver data needed"],
        "external": d["Key external data / tools"],
        "models": d["Candidate model(s) / tool(s)"],
        "status": d["Status"],
        "verdict": d["Empirical verdict (Q-Mammal where applicable)"],
        "gap": d["Gap -> build?"],
        "repPrompts": d.get("Representative prompts", ""),
    })

# prompt counts + disease mix per capability
per_cap = Counter()
dis_per_cap = defaultdict(Counter)
dis_total = Counter()
for row in pr_ws.iter_rows(min_row=2, values_only=True):
    pid, text, cat, cid, carea, dis = row
    per_cap[cid] += 1
    dis_per_cap[cid][dis] += 1
    dis_total[dis] += 1
for c in caps:
    c["promptCount"] = per_cap.get(c["id"], 0)
    c["topDiseases"] = dis_per_cap[c["id"]].most_common(4)

# ---- hardcoded from integration_map.md (not in xlsx) ----
layers = {
  "Internal": {
    "blurb": "Quiver's unique functional data — the moat. Novel target signals nobody else has.",
    "sources": [["Quiver EP-CRISPR Atlas",104,1],["GenomicsDB",27,2],["Sapphire Embedding Engine",9,3]],
  },
  "Context": {
    "blurb": "External data Quiver can't know — gates go/no-go. “Great pain target but causes cancer → no-go.”",
    "sources": [["GTEx",45,1],["ClinVar",42,1],["OMIM",38,1],["DisGeNET",35,1],["HPO",29,2],["TCGA",27,2],["gnomAD",11,3],["Allen Brain Atlas",8,3],["Open Targets",7,3],["Cortellis",7,3]],
  },
  "Predictivity": {
    "blurb": "Independent corroboration → re-ranks hits. The #7 → #1 boost.",
    "sources": [["STRING",45,1],["Reactome",39,1],["BioGRID",34,1],["LINCS L1000",29,2],["Connectivity Map",27,2],["Expression Atlas",27,2],["KEGG",10,3],["GWAS Catalog",7,3]],
  },
  "Reference": {
    "blurb": "Identifiers, structures, libraries, algorithms — the plumbing every layer depends on.",
    "sources": [["DrugBank",55,1],["UniProt",45,1],["ChEMBL",17,2],["PDB / RCSB",13,3],["PubChem",11,3],["PubMed",11,3],["UCSC Genome Browser",11,3],["RNAfold",10,3],["NUPACK",9,3],["RDKit",9,3],["BLAST",8,3],["Scanpy / Seurat",8,3],["FAISS",7,3]],
  },
}

methodology = [
  {"k":"personas","n":"59","t":"Personas","d":"Fake-but-real-mandate execs (Pharma SVP, BD, Biotech CSO, VC GP) — who asks Sapphire, and with what philosophy."},
  {"k":"prompts","n":"299","t":"Prompts","d":"The questions each persona would put to Sapphire — the demand side, across ~25 categories."},
  {"k":"pipelines","n":"399","t":"Pipelines","d":"Each prompt expanded into a uniform 6-stage decomposition (inputs → tools → sub-prompts → outputs). 299 Sapphire + 100 Angelini."},
  {"k":"frequency","n":"—","t":"Tool frequency","d":"Aggregate every pipeline's tool calls into a ranked list — the prioritized integration roadmap."},
  {"k":"map","n":"16","t":"Capability map","d":"Cluster prompts into 16 capability areas; map each to candidate models + empirical status."},
  {"k":"gaps","n":"2","t":"Gaps → build","d":"Where no off-the-shelf model exists (CAP-04 ASO design, CAP-15 expert judgment) → Quiver builds: curated corpus, expert agent, or data fine-tune."},
]

# Sample query for the system-flow animation (the real Nav1.8 SCN11A #7->#1 run; matches the Console)
sampleQuery = {
  "query": "Prioritize novel analgesic targets in the Nav1.8 (SCN10A) network for a systemic neuropathic-pain program.",
  "tier": "Unified Orchestration (400–2000 ms)",
  "stages": [
    {"id":"L1","name":"Internal latent","kind":"internal",
     "detail":"Quiver EP-CRISPR fused embedding + DrugReflector rank the candidates.",
     "result":"SCN11A/Nav1.9 lands at #7 — its ultra-slow persistent current is under-resolved by the optical-EP assay."},
    {"id":"L2","name":"Context gate","kind":"gate",
     "detail":"EMET drug-safety / FDA-memory — a veto channel only (demote / kill, never promote).",
     "result":"Vetoes CACNA2D1 (black-box respiratory depression) and NGF (RPOA, NDA rejected 2021). SCN11A passes — SAFE class."},
    {"id":"L3","name":"Predictivity boost","kind":"boost",
     "detail":"Independent corroboration: GWAS / Mendelian genetics, STRING PPI with the disease gene, academic functional screens, LINCS signature.",
     "result":"SCN11A corroborated — FEPS3 Mendelian genetics + the panel's only SAFE class → re-ranked #7 → #1."},
    {"id":"OUT","name":"Answer + execution plan","kind":"out",
     "detail":"Calibrated uncertainty gate. Confident → emit ranked hits + which embeddings/sources moved each rank. Uncertain → abstain & propose the experiment.",
     "result":"SCN11A/Nav1.9 #1 with provenance; one gate remains — cardiac Nav1.5 selectivity. Biology: HIGH."},
  ],
  "tiers": [
    ["Direct Run","<100 ms","internal latent similarity only"],
    ["Atomic Fusion","200–500 ms","the re-ranking cascade (common case)"],
    ["Unified Orchestration","400–2000 ms","planner + multi-agent panel over the metagraph"],
  ],
}

data = {
  "meta": {"personas":59,"prompts":299,"pipelines":399,"capabilities":16,"atlasFreq":104,
           "kgNodes":"1.8M","kgEdges":"17.9M","papers":"29k"},
  "capabilities": caps,
  "layers": layers,
  "methodology": methodology,
  "sampleQuery": sampleQuery,
}

os.makedirs(os.path.dirname(OUT), exist_ok=True)
with open(OUT, "w", encoding="utf-8") as f:
    f.write("// Auto-generated from capability_map.xlsx by _build/build_site_data.py\n")
    f.write("window.SAPPHIRE = ")
    json.dump(data, f, ensure_ascii=False, indent=2)
    f.write(";\n")
print("wrote", OUT)
print("capabilities:", len(caps), "| prompt counts sum:", sum(per_cap.values()))
