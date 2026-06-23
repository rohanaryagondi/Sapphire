# -*- coding: utf-8 -*-
"""Build capability_map.xlsx: Capabilities sheet (16 areas) + Prompts sheet (299)."""
import re, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

CHECKLIST = r"C:\Users\rohan.gondi\Desktop\Sapphire\extracted\Sapphire Prompt Work_Feb 2026\Sapphire_Pipeline_Master_Checklist.md"
# Output the generated workbook into docs/foundation/ (repo-relative; was a Windows abs path).
OUT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "docs", "foundation", "capability_map.xlsx")

# ---------- parse the 299 prompts ----------
prompts = []          # (id, text, category)
cat = None
with open(CHECKLIST, encoding="utf-8") as fh:
    for line in fh:
        m = re.match(r"^##\s+(.*?)\s*\(\d+/\d+", line)
        if m:
            cat = re.sub(r"\s*\(.*$", "", m.group(1)).strip(); continue
        m = re.match(r"^- \[.\] \*\*(\d+)\*\*\s*[—-]\s*(.*)$", line.strip())
        if m:
            prompts.append((int(m.group(1)), m.group(2).strip(), cat or ""))
prompts.sort()

# ---------- capability assignment ----------
def cap_of(pid, text):
    t = text.lower()
    # explicit per-prompt overrides first
    if pid in (6,): return "CAP-01"
    if pid in (7,): return "CAP-13"
    if pid in (9,): return "CAP-08"
    if pid in (30,): return "CAP-02"
    if pid in (70,135): return "CAP-13"
    if pid in (71,): return "CAP-02"
    if pid in (72,): return "CAP-08"
    if pid in (73,): return "CAP-14"
    if pid in (74,): return "CAP-15"
    if pid in (99,100): return "CAP-03"
    if pid in (102,): return "CAP-08"
    if pid in (143,): return "CAP-13"
    if pid in (144,): return "CAP-14"
    if pid in (273,): return "CAP-14"
    # range-based defaults (master-checklist sections)
    if 1   <= pid <= 9:   return "CAP-04"
    if pid in (10,11,12) or 31 <= pid <= 41: return "CAP-01"
    if 13  <= pid <= 20:  return "CAP-02"
    if 42  <= pid <= 43:  return "CAP-02"
    if 21  <= pid <= 29 or pid == 44: return "CAP-04"
    if 45  <= pid <= 54:  return "CAP-03"
    if 55  <= pid <= 59:  return "CAP-09"
    if 60  <= pid <= 64:  return "CAP-10"
    if 65  <= pid <= 69:  return "CAP-14"
    if 75  <= pid <= 84:  return "CAP-05"
    if 85  <= pid <= 94:  return "CAP-06"
    if 95  <= pid <= 104: return "CAP-07"
    if 105 <= pid <= 114: return "CAP-08"
    if 115 <= pid <= 124: return "CAP-09"
    if 125 <= pid <= 134: return "CAP-04" if "aso" in t else "CAP-02"
    if 136 <= pid <= 142: return "CAP-13"
    if 145 <= pid <= 154: return "CAP-14"
    if 155 <= pid <= 164: return "CAP-16"
    if 165 <= pid <= 174: return "CAP-15"
    # disease application batches 175-274: route by intent keyword
    if 175 <= pid <= 274:
        if "aso" in t or "antisense" in t or "gapmer" in t or "splice" in t: return "CAP-04"
        if "combination" in t or "combinatorial" in t or "polypharm" in t or "dual-mechanism" in t: return "CAP-09"
        if "budget" in t or "prioritize 3" in t or "portfolio" in t: return "CAP-14"
        if "biomarker" in t: return "CAP-10"
        return "CAP-02"
    if 275 <= pid <= 299: return "CAP-04"
    return "CAP-02"

def disease_of(text):
    t = text.lower()
    rules = [
        ("Pain / channelopathy", ["nav1.7","nav1.8","scn9a","scn8a","pain","neuropath","nocicep","dorsal root"]),
        ("Epilepsy / DEE",       ["epilep","seizure","dee","dravet","scn1a","scn2a","kcnq2","stxbp1","cdkl5","depdc5","slc6a1"]),
        ("TSC / mTOR",           ["tsc1","tsc2"," tsc","mtor","rapalog"]),
        ("Rett / MECP2",         ["mecp2","rett","cdkl5","foxg1"]),
        ("ASD / NDD",            ["asd","autism","shank3","syngap","dup15q","angelman","ube3a","grin2b","cacna1a","cntnap2","phelan","kif1a","fragile x","neurodevelopmental"]),
        ("Alzheimer's",          ["alzheim","tauopath","tau ","apoe"," ad ","early ad"]),
        ("Parkinson's",          ["parkinson","synuclein","dopaminergic"]),
        ("ALS",                  ["als","motor neuron"]),
        ("Huntington's",         ["huntington","htt"]),
        ("FTD / dementia",       ["ftd","mapt","dementia","frontotemporal"]),
        ("Schizophrenia",        ["schizophren","psychos","antipsychotic","anti-psychotic"]),
        ("Depression / mood",    ["depress","antidepress","ssri","mood","bipolar","lithium"]),
        ("Psychiatry (other)",   ["ptsd","ocd","stress circuit","sleep"]),
    ]
    for label, kws in rules:
        if any(k in t for k in kws): return label
    return "Cross-disease / platform"

# ---------- capability area metadata (Sheet 1) ----------
CAPS = [
 ("CAP-01","Functional similarity / embedding & clustering","EP-signature proximity, antipodal/proximity, pathway reconstruction, KEGG clustering.","Internal","EP-CRISPR Atlas","STRING, Reactome, KEGG","Quiver encoder (native); MAMMAL (0.92), ESM-2-650M, ESM-C, SaProt, Ankh","Tested","MAMMAL NN-recall 0.92 on protein families, ties ESM-2-650M (0.75) on the 40-gene CRISPR-N panel; PROTON lost (0.49). Embedding/clustering is real & useful off-the-shelf. [Q-Mammal]","—"),
 ("CAP-02","Target discovery & prioritization","Rank targets, antipodal-to-disease, convergent/common nodes across diseases.","Internal","EP-CRISPR Atlas","OpenTargets, DisGeNET, GWAS Catalog","Quiver DrugReflector (native); Open Targets + L2G, PandaOmics, TXGNN, PINNACLE","Native","Core Quiver function; external genetics/competition data re-ranks. Not an off-the-shelf-model question.","Integrate CAP-06/CAP-12 for boosting"),
 ("CAP-03","Drug-target binding / DTI / ligandability","Single-target binder vs decoy triage, small-molecule rescue match, repurposing, ligandability.","Internal + Predictivity","EP-CRISPR Atlas","ChEMBL, BindingDB, DrugBank, RCSB PDB","Boltz-2 (mTOR 1.0/Nav 0.71), Chai-1, AF3, DeepPurpose; MAMMAL/ConPLex = chance on Nav","Tested","Off-the-shelf single-target triage ~= chance (MAMMAL Nav1.8 0.43, ConPLex 0.39). Boltz-2 split: mTOR AUROC 1.00, Nav1.8 0.71 (marginal, n=28). No off-the-shelf Nav oracle. [Q-Mammal]","Quiver-data fine-tune for Nav-like triage (the lever)"),
 ("CAP-04","ASO design & sequence generation","Knockdown / allele-specific / splice-modulating ASOs, chemistry ranking, exact sequences.","Internal","EP-CRISPR Atlas + transcriptomics","gnomAD, ClinVar, RNAfold/ViennaRNA, NUPACK","RNAfold, NUPACK, SpliceAI/Pangolin, eSkip-Finder; OligoAI/ASOptimizer (preprint); no mature oracle","Gap","MAMMAL public artifact is a span-infiller (no usable de-novo design). ASO sequence design needs a dedicated stack. [Q-Mammal]","Build / curate ASO-design toolchain"),
 ("CAP-05","Mechanism disambiguation","Synaptic vs intrinsic excitability, upstream/downstream, disease-modifying vs symptomatic.","Internal","EP-CRISPR Atlas (EP)","Reactome, KEGG","Quiver EP-assay decomposition (native)","Native","Quiver-unique: separates synaptic vs intrinsic signatures directly from electrophysiology.","—"),
 ("CAP-06","Genetics <-> function integration","GWAS support, ClinVar mapping onto perturbation space, protective-variant simulation.","Predictivity","EP-CRISPR Atlas","GWAS Catalog, ClinVar, OMIM, DisGeNET","AlphaMissense, ESM-variant, VEP, Open Targets Genetics/L2G + KG join (Neo4j)","Untested","The boosting layer James described ('re-rank #7 to #1'): independent genetic corroboration of functional hits.","Stand up KG join"),
 ("CAP-07","BBB / PK-PD / druggability","BBB penetrance, required CNS exposure, metabolic liability.","Context","none / Atlas","SwissADME, ADMET-AI, B3DB","ADMET-AI (preferred), ADMETLab 3.0, SwissADME, pkCSM; MAMMAL BBBP (FP-biased)","Tested","MAMMAL BBBP AUROC 0.97 but false-positive biased -> soft positive only; ADMET-AI calibrated endpoints preferred. [Q-Mammal]","—"),
 ("CAP-08","Toxicity & safety prediction","Seizure/immunogenicity/ADMET risk, contraindications, the 'great target but causes cancer' check.","Context","none","ToxCast, Tox21, ADMET-AI, FAERS, VigiBase","ADMET-AI (DILI 0.83), ProTox-III, DeepTox, hERG ensembles; MAMMAL ClinTox unusable; seizure-liability = gap","Tested","ADMET-AI earns the tox-gate slot; MAMMAL ClinTox is memorized (0% sensitivity to external toxics) - do not gate on it. [Q-Mammal]","—"),
 ("CAP-09","Combination & network strategy","Synergistic pairs, dual-target profiles, hub genes that collapse multiple clusters.","Internal","EP-CRISPR Atlas","STRING, BioGRID, Reactome","Quiver combination-rescue (native) + network propagation","Native / Untested","Quiver functional combinations native; network topology from PPI/pathway DBs adds candidates.","—"),
 ("CAP-10","Biomarker & translational","Fluid biomarkers aligned to EP phenotype, EEG correlation, symptom-domain mapping.","Predictivity","EP-CRISPR Atlas","GEO, Expression Atlas, Human Protein Atlas","Correlation analysis + literature","Untested","—","—"),
 ("CAP-11","Variant -> disease / prevalence / genetic epidemiology","Is gene X associated with disease Y; pathogenic-variant counts; patient population size.","Context","none","ClinVar, HGMD, gnomAD, OMIM, HPO, Orphanet","Knowledge-graph lookup (Neo4j, 1.8M nodes / 17.9M edges)","Untested","James' 'Amit question' set: a knowledge-graph/dataset capability, not a model. Largely served by the existing Sapphire v3 Neo4j graph.","—"),
 ("CAP-12","Protein-protein interaction / pathway membership","Is gene in pathway X; PPI partners; cross-assay corroboration for re-ranking.","Predictivity","none","STRING, BioGRID, Reactome, KEGG","Graph / KG query","Untested","Boosting layer: independent assay corroboration (the 'appeared in two assays -> highly compelling' logic).","—"),
 ("CAP-13","Competitive & commercial intelligence","Active programs by phase, white-space, pricing, peak-sales, exclusivity.","Context","none","ClinicalTrials.gov, Cortellis, GlobalData, IQVIA, patents","LLM + curated feeds (Emit-style agentic web+DB)","Untested","This is the layer Emit/Sapphire orchestration must cover; test against the disease-MoA white-space prompts.","Curated competitive-intel corpus"),
 ("CAP-14","Portfolio & capital allocation / financial optimization","rNPV, $100M/$10B allocation, kill/fast-track decisions, peak-sales modeling.","Meta","none","Tufts CSDD, IQVIA, GlobalData; Gurobi/CPLEX solvers","LLM reasoning + optimization solvers","Untested","Pipelines already specify Monte-Carlo / ILP solvers; mostly LLM-orchestrated reasoning over market data.","—"),
 ("CAP-15","Expert judgment / strategic reasoning","'If you were CSO at Lilly...', regulatory strategy, clinical-trial design, franchise calls.","Meta","none","Public expert content (blogs, podcasts, talks, posts), regulatory precedent","EXPERT-AGENT (to build) - the '$50k Pfizer expert from public posts' idea","Gap","No off-the-shelf model. James' headline build: emulate a CNS regulatory/clinical expert from public output; the stock-sentiment-bot pattern applied to biology.","BUILD expert-agent corpus + retrieval"),
 ("CAP-16","AI self-reflection / uncertainty quantification","Where is data insufficient, confidence intervals, replication strength, model-drift sensitivity.","Meta","EP-CRISPR Atlas (replication)","—","Conformal (MAPIE/TorchCP), deep ensembles, abstention; Sapphire v3 plan layer","Partial / Native","Maps to Sapphire v3's transparent execution plans (which embeddings contributed, where data contradicts).","—"),
]
CAP_HEADERS = ["ID","Capability area","Description","Layer","Quiver data needed","Key external data / tools","Candidate model(s) / tool(s)","Status","Empirical verdict (Q-Mammal where applicable)","Gap -> build?"]

# rep prompt ids per capability, from the mapping
from collections import defaultdict
rep = defaultdict(list)
rows_prompts = []
for pid, text, category in prompts:
    cid = cap_of(pid, text)
    dis = disease_of(text)
    rep[cid].append(pid)
    rows_prompts.append((pid, text, category, cid, dis))

def fmt_ids(ids):
    ids = sorted(ids); out=[]; i=0
    while i < len(ids):
        j=i
        while j+1 < len(ids) and ids[j+1]==ids[j]+1: j+=1
        out.append(f"{ids[i]:03d}" if i==j else f"{ids[i]:03d}-{ids[j]:03d}")
        i=j+1
    return ", ".join(out)

# ---------- styling ----------
ARIAL = "Arial"
HFILL = PatternFill("solid", fgColor="1F3864")   # Quiver-ish deep blue
HFONT = Font(name=ARIAL, bold=True, color="FFFFFF", size=11)
BODY  = Font(name=ARIAL, size=10)
IDFONT= Font(name=ARIAL, bold=True, size=10)
TOP   = Alignment(vertical="top", wrap_text=True)
thin  = Side(style="thin", color="D9D9D9")
BORD  = Border(left=thin,right=thin,top=thin,bottom=thin)
LAYER_FILL = {
 "Internal":"DDEBF7","Internal + Predictivity":"DDEBF7","Internal + Predictivity ":"DDEBF7",
 "Predictivity":"E2EFDA","Context":"FCE4D6","Meta":"EDEDED",
}
STATUS_FILL = {"Tested":"C6EFCE","Native":"D9E1F2","Native / Untested":"D9E1F2","Partial / Native":"D9E1F2","Untested":"FFF2CC","Gap":"FFC7CE"}

def style_header(ws, ncols):
    for c in range(1, ncols+1):
        cell = ws.cell(1, c); cell.fill=HFILL; cell.font=HFONT
        cell.alignment=Alignment(vertical="center", wrap_text=True); cell.border=BORD
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

wb = Workbook()

# ----- Sheet 1: Capabilities -----
ws = wb.active; ws.title = "Capabilities"
ws.append(CAP_HEADERS + ["Representative prompts"])
# CAP-11/CAP-12 are James' verbal additions (the meeting), not present in the 299-prompt corpus
VERBAL = {
 "CAP-11":"(none in the 299 - James' verbal add re: prevalence/variant-disease; nearest: 076, 089, 129)",
 "CAP-12":"(none in the 299 - James' verbal add re: PPI; nearest functional analogs: 033, 036-040)",
}
for r in CAPS:
    cid = r[0]
    reps = fmt_ids(rep[cid]) if rep[cid] else VERBAL.get(cid, "")
    ws.append(list(r) + [reps])
widths = [8,30,40,14,18,28,32,12,52,30,26]
for i,w in enumerate(widths, start=1):
    ws.column_dimensions[ws.cell(1,i).column_letter].width = w
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.font=BODY; cell.alignment=TOP; cell.border=BORD
    row[0].font=IDFONT
    lf = LAYER_FILL.get(row[3].value.strip() if row[3].value else "")
    if lf: row[3].fill=PatternFill("solid", fgColor=lf)
    sf = STATUS_FILL.get(row[7].value)
    if sf: row[7].fill=PatternFill("solid", fgColor=sf)
style_header(ws, len(CAP_HEADERS)+1)

# ----- Sheet 2: Prompts -----
ws2 = wb.create_sheet("Prompts")
ws2.append(["ID","Prompt","Original category","Capability ID","Capability area","Disease area"])
cap_name = {c[0]:c[1] for c in CAPS}
for pid, text, category, cid, dis in rows_prompts:
    ws2.append([f"{pid:03d}", text, category, cid, cap_name[cid], dis])
for i,w in enumerate([7,80,34,12,30,26], start=1):
    ws2.column_dimensions[ws2.cell(1,i).column_letter].width = w
for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
    for cell in row:
        cell.font=BODY; cell.alignment=TOP; cell.border=BORD
    row[0].font=IDFONT
style_header(ws2, 6)

# ----- Sheet 3: README/legend -----
ws3 = wb.create_sheet("How to read", 0)  # put first
notes = [
 ("Sapphire Capability Map", 14, True),
 ("", 10, False),
 ("Built from James' Feb-2026 Sapphire prompt corpus (299 prompts, 59 personas, 399 pipelines, tool-frequency analysis) + the Q-Mammal empirical model evaluation.", 10, False),
 ("", 10, False),
 ("SHEETS", 11, True),
 ("- Capabilities: 16 capability areas (the strategy view). One row per underlying capability Sapphire must serve.", 10, False),
 ("- Prompts: all 299 customer prompts, each mapped to a capability area + disease area (the backing detail).", 10, False),
 ("", 10, False),
 ("LAYER (James' 3-layer data vision)", 11, True),
 ("- Internal  = Quiver's unique EP-CRISPR / V1-T data; the moat.", 10, False),
 ("- Context   = external data Quiver cannot know (safety, prevalence, competition) that gates go/no-go.", 10, False),
 ("- Predictivity = independent corroboration (genetics, PPI, pathway) that re-ranks/boosts hits.", 10, False),
 ("- Meta      = reasoning/judgment layers (financial optimization, expert judgment, uncertainty).", 10, False),
 ("", 10, False),
 ("STATUS", 11, True),
 ("- Tested = empirical result exists in Q-Mammal.  Native = core Quiver function.  Untested = no eval yet.  Gap = no off-the-shelf solution -> see 'Gap -> build?'.", 10, False),
 ("", 10, False),
 ("HONESTY RULE: Status/verdict reflect only what Q-Mammal established. Everything else is 'Untested' or 'Gap' - no invented model performance.", 10, False),
 ("", 10, False),
 ("FULL MODEL DETAIL: the 'Candidate model(s)' column is a summary. See docs/foundation/model_landscape.md for the full per-capability landscape (3-6 named models each, with maturity + proven/paper-claim flags).", 10, False),
]
for i,(txt,sz,bold) in enumerate(notes, start=1):
    c = ws3.cell(i,1, txt); c.font=Font(name=ARIAL, size=sz, bold=bold, color=("1F3864" if bold and sz>=11 else "000000")); c.alignment=Alignment(wrap_text=True, vertical="top")
ws3.column_dimensions["A"].width = 130

wb.save(OUT)
print("saved", OUT)
print("prompts parsed:", len(prompts))
from collections import Counter
cc = Counter(cap_of(p,t) for p,t,_ in prompts)
for k in sorted(cc): print(k, cc[k], cap_name[k])
print("unmapped diseases sample:", Counter(disease_of(t) for _,t,_ in prompts))
